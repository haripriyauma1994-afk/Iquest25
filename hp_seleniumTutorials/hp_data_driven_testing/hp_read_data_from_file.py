'''
Created on 03-Jan-2026

@author: ABC
'''
import openpyxl
import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By

#1 complete workbook
filename = r"C:\Users\ABC\Documents\DOT_2026_python.xlsx" #complete workbook
my_workbook = openpyxl.load_workbook(filename) #filename
#2.sheet
active_sheet = my_workbook["Sheet1"] # loads the sheet by sheet name

#3.cell module
cell_value_username = active_sheet.cell(2, 1).value
print(cell_value_username)

cell_value_password = active_sheet.cell(2, 2).value
print(cell_value_password)

class TestSetUpTearDown(unittest.TestCase):


    def setUp(self):
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)
        options.add_argument("start-maximized")
        self.driver = webdriver.Chrome(options)
        self.driver.implicitly_wait(5)
        self.driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")


    def tearDown(self):
        pass


    def test_navigation_to_orangehrm_loginpage(self):
        expected_url= ("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        current_page_url = self.driver.current_url
        self.assertEqual(expected_url, current_page_url, "current page url is different from expected url")
        
    
    def test_login_to_orangehrm(self):
        self.driver.find_element(By.NAME, "username").send_keys("Admin")
        self.driver.find_element(By.NAME, "password").send_keys("admin123")
        
        login_button = self.driver.find_element(By.XPATH,"//button[@type='submit']")
        login_button.click()
        
        expected_url= ("https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index")
        current_page_url = self.driver.current_url
        self.assertEqual(expected_url, current_page_url, "current page url is different from expected url")


if __name__ == "__main__":
    #import sys;sys.argv = ['', 'Test.testName']
    unittest.main()